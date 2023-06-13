import openpyxl, time
from openpyxl import Workbook

# listas para salvar os identificadores únicos e não ter cadastro duplicados   
id_aluno = []
id_livro = []
id_emprestimos = []

# dicionarios para vincular identificadores únicos aos atributos 
info_alunos = {}
info_livros = {}

# dicionarios para emprestimos
dic_emprestimos = {}


# criação da base de dados
# verifica se a planilha já existe e cria o arquivo caso não exista
try:
    database_ids = openpyxl.load_workbook("biblioteca.xlsx")

except FileNotFoundError:
    database_ids = Workbook()
    database_ids.save("biblioteca.xlsx")

# cria uma sheet chamada alunos na base de dados caso não exista
if "alunos"  in database_ids.sheetnames:
    planilha_alunos = database_ids["alunos"]
else:
    planilha_alunos = database_ids.create_sheet("alunos")
    planilha_alunos.title = "alunos"
    database_ids.save("biblioteca.xlsx")

# for para exportar os identificadores dos alunos para a lista 
for celula in planilha_alunos["A"]:
    id_aluno.append(celula.value)

# o primeiro for cria uma chave de dicionáro para cada identificador único do aluno com uma lista vazia
# o segundo for itera sobre as colunas da planilha e adiciona os dados(das linhas) na lista vinculando o 
# identificador único a cada conjunto de atribunos ex: nome, idade, turno...

for a in range(len(id_aluno)):                              
    info_alunos[id_aluno[a]] = []                           
    for coluna in planilha_alunos.iter_cols():              
        info_alunos[id_aluno[a]].append(coluna[a].value)    

if "livros"  in database_ids.sheetnames:
    planilha_livros = database_ids["livros"]
else:
    planilha_livros = database_ids.create_sheet("livros")
    planilha_livros.title = "livros"
    database_ids.save("biblioteca.xlsx")
  
for celula in planilha_livros["A"]:
    id_livro.append(celula.value)

# o primeiro for cria uma chave de dicionáro para cada identificador único de livro com uma lista vazia
# o segundo for itera sobre as colunas da planilha e adiciona os dados(das linhas) na lista vinculando o 
# identificador único a cada conjunto de atribunos ex, titulo, autor, editora...

for a in range(len(id_livro)):                              
    info_livros[id_livro[a]] = []                           
    for coluna in planilha_livros.iter_cols():           
        info_livros[id_livro[a]].append(coluna[a].value) 


# cria uma sheet chamada emprestimos na base de dados caso não exista
if "emprestimos"  in database_ids.sheetnames:
    planilha_emprestimos = database_ids["emprestimos"]
else:
    planilha_emprestimos = database_ids.create_sheet("emprestimos")
    planilha_emprstimos.title = "emprestimos"
    database_ids.save("biblioteca.xlsx")

for celula in planilha_emprestimos["A"]:
    id_emprestimos.append(celula.value)

for a in range(len(id_emprestimos)):                              
    dic_emprestimos[id_emprestimos[a]] = []                           
    for coluna in planilha_emprestimos.iter_cols():              
        dic_emprestimos[id_emprestimos[a]].append(coluna[a].value)


# funções para cadastro de alunos e livros
def cadastra_aluno():
    # criação de verificador unico    
    verificador = len(id_aluno)
    
    if verificador not in id_aluno:
        id_aluno.append(verificador)
        
    else:
        verificador += 1
    print("ID =",verificador)
    nome = input("Nome do Aluno: ")
    

    serie = input("Série: ")
    turno = input("Turno: ")
    idade = int(input("Idade: "))
    print("Contato: 00 0 0000 0000")
    contato = input()
    print("Endereço: Rua, Bairro, Número.:")
    endereco = input()

    # atualizando dicionário
    info_alunos[verificador] = f"Nome = {nome.capitalize()}, Série = {serie}, Turno = {turno}, Idade = {idade}, Contato = {contato}, Endereço = {endereco.capitalize()}"

    print()
    print("Os dados do aluno cadastrados são: ")
    print(info_alunos[verificador])
    print()

    proxima_linha_aluno = planilha_alunos.max_row + 1

    # atualizando a planilha
    planilha_alunos[f"A{proxima_linha_aluno}"] = verificador
    planilha_alunos[f"B{proxima_linha_aluno}"] = nome.capitalize()
    planilha_alunos[f"C{proxima_linha_aluno}"] = serie
    planilha_alunos[f"D{proxima_linha_aluno}"] = turno.capitalize()
    planilha_alunos[f"E{proxima_linha_aluno}"] = idade
    planilha_alunos[f"F{proxima_linha_aluno}"] = contato
    planilha_alunos[f"G{proxima_linha_aluno}"] = endereco.capitalize()

    # salvando dados
    database_ids.save("biblioteca.xlsx")

def cadastra_livro():    
    
    titulo_livro = input("Digite o título do livro: ")
    genero = input("Digite o gênero do livro: ")
    autor = input("Digite o Autor: ")
    editora = input("Digite a Editora: ")
    qtd = input("Digite a quantidade: ")
    # verificando se a quantidade é um valor numérico
    while True:
        if qtd.isdigit() == False:
            qtd = input("Digite um valor válido, um número: ")
        else:
            break
    numeracao = input("Digite a numeração: ")
    # verificando se a numeração do livro é um valor numérico
    while True:
        if numeracao.isdigit():
            while True:
                if int(numeracao) not in id_livro and numeracao.isdigit():
                    break
                else:
                    print("Livro já cadastrado ou numeração inválida.")
                    numeracao = input("Digite uma numeração válida: ")
            break
        else:
            print("Digite uma numeração válida.")

    numeracao = int(numeracao)

    proxima_linha_livro = planilha_livros.max_row + 1
    # atualizando a planilha
    planilha_livros[f"A{proxima_linha_livro}"] = numeracao
    planilha_livros[f"B{proxima_linha_livro}"] = titulo_livro.capitalize()
    planilha_livros[f"C{proxima_linha_livro}"] = genero.capitalize()
    planilha_livros[f"D{proxima_linha_livro}"] = autor.capitalize()
    planilha_livros[f"E{proxima_linha_livro}"] = editora.capitalize()
    planilha_livros[f"F{proxima_linha_livro}"] = int(qtd)

    # atualizando dicionário
    info_livros[numeracao] = f"Título = {titulo_livro.capitalize()}, Gênero = {genero.capitalize()}, Autor = {autor.capitalize()},Editora =  {editora.capitalize()}, Quantidade = {qtd}, Numeração = {numeracao}"
    id_livro.append(numeracao) 

    print()
    print("As informações do livro cadasrtado são:")  
    print("NUMERAÇÃO: ", numeracao) 
    print(info_livros[numeracao])
    print()
    database_ids.save("biblioteca.xlsx")


# funções para alterações de cadastro
def altera_livro():
    # verificação de existencia do livro
    while True:
        numeracao = int(input("Digite a numeração do livro que quer alterar: "))
        if numeracao not in id_livro:
            print("Livro não cadastrado ou numeração inválida, revise e digite uma numeração válida.")
            continue
        else:
            print("O livro que vai ser alterado é:")
            print(info_livros[numeracao])
            print("Digite as alterações:")

            # cadastro de alterações
            titulo_livro = input("Digite o título do livro: ")
            genero = input("Digite o gênero do livro: ")
            autor = input("Digite o Autor: ")
            editora = input("Digite a Editora: ")
            qtd = input("Digite a quantidade: ")
            while True:
                if qtd.isdigit() == False:
                    qtd = input("Digite um valor válido, um número: ")
                else:
                    qtd = int(qtd)
                    break
            
            # atualizando dados no dicionário
            info_livros[numeracao] = f"Título = {titulo_livro.capitalize()}, Gênero = {genero.capitalize()}, Autor = {autor.capitalize()},Editora =  {editora.capitalize()}, Quantidade = {qtd}"

            linha_livro = id_livro.index(numeracao) + 1

            planilha_livros[f"B{linha_livro}"] = titulo_livro.capitalize()
            planilha_livros[f"C{linha_livro}"] = genero.capitalize()
            planilha_livros[f"D{linha_livro}"] = autor.capitalize()
            planilha_livros[f"E{linha_livro}"] = editora.capitalize()
            planilha_livros[f"F{linha_livro}"] = qtd

            break
                
def altera_aluno():
    while True:
        verificador = int(input("Digite o ID do Aluno que quer alterar: "))
        if verificador not in id_aluno:
            print("Aluno não cadastrado ou numeração inválida, revise e digite uma numeração válida.")
            continue
        else:
            print("O cadastro que vai ser alterado é:")
            print(info_alunos[verificador])
            print("Digite as alterações:")

            nome = input("Nome do Aluno: ")
            serie = str(input("Série: "))
            turno = input("Turno: ")
            idade = int(input("Idade: "))
            print("Contato: 00 0 0000 0000")
            contato = input()
            print("Endereço: Rua, Bairro, Número.:")
            endereco = input()

            # atualizando dicionário
            info_alunos[verificador] = f"Nome = {nome.capitalize()}, Série = {serie}, Turno = {turno}, Idade = {idade}, Contato = {contato}, Endereço = {endereco.capitalize()}"

            # atualizando planilha

            planilha_alunos[f"B{verificador + 1}"] = nome.capitalize()
            planilha_alunos[f"C{verificador + 1}"] = serie
            planilha_alunos[f"D{verificador + 1}"] = turno.capitalize()
            planilha_alunos[f"E{verificador + 1}"] = idade
            planilha_alunos[f"F{verificador + 1}"] = contato
            planilha_alunos[f"G{verificador + 1}"] = endereco.capitalize()

            # salvando dados
            database_ids.save("biblioteca.xlsx")
            break


# funções para empréstimo e devolução de livros

def emprestimo():
    global info_alunos

    r = input("Tem conhecimento do ID do aluno? [S]im [N]ão: ")
    if r in "nN":
        nome = input("Digite o nome do aluno: ")

        contador = 0
        print(info_alunos["ID"])   
        for chave in info_alunos:    
            if nome in info_alunos[chave][1]:  
                print(info_alunos[chave])
                contador += 1
        if contador == 0:
            print("Aluno não encontrado.")
    
    while True:
        key_aluno = int(input("Digite o ID do aluno: "))

        if key_aluno not in info_alunos:
            print("ID não encontrado, digite um ID válido.")
            continue

        else:

            livro = input("Digite o Livro que será emprestado: ")
            devo = str(input("Digite a data para devolução(Dia/Mês): "))
            chave = str(key_aluno)+devo.replace("/", "")
            
            while chave  in dic_emprestimos:
                print("O aluno já tem um livro para entregar nessa data, adicione mais um dia na data de entrega para poder emprestar mais um livro.")
                devo = str(input("Digite a data para devolução(Dia/Mês): "))
                chave = str(key_aluno)+devo.replace("/", "")
                
            else:
                dic_emprestimos[chave] = info_alunos[key_aluno], livro, devo
                print(dic_emprestimos[chave])
                break

    linha = planilha_emprestimos.max_row + 1
    # atualizando a planilha
    planilha_emprestimos[f"A{linha}"] = str(chave)
    planilha_emprestimos[f"B{linha}"] = str(info_alunos[key_aluno])
    planilha_emprestimos[f"C{linha}"] = livro
    planilha_emprestimos[f"D{linha}"] = devo
    database_ids.save("biblioteca.xlsx")

def devolucao():
    global dic_emprestimos
    

    r = input("Tem conhecimento do ID do aluno? [S]im [N]ão: ")
    if r in "nN":
        nome = input("digite o nome do aluno: ")
        contador = 0
        print(f"Registros de {nome} nas pendências de devolução.")
        print(dic_emprestimos["ID EMPRESTIMO"])  
        for chave in dic_emprestimos:    
            if nome in dic_emprestimos[chave][1]:  
                print(dic_emprestimos[chave])
                contador += 1
        if contador == 0:
            print("Aluno não encontrado.")
            
    while True:
        idaluno = input("Digite o ID do aluno: ")
        devo = str(input("DIGITE A DATA QUE FOI PROGRAMADA PARA ENTREGA(DD/MM): "))      
        chave = idaluno+devo.replace("/", "")
        if chave in dic_emprestimos:
            
            linha = 1
            for celula in planilha_emprestimos["A"]:
                if celula.value == chave:
                    planilha_emprestimos.delete_rows(linha)
                    break
                linha +=1
                
            dic_emprestimos.pop(chave)
            database_ids.save("biblioteca.xlsx")
            print()
            print("Devolução realizada.")
            print()
            break
        else:
            print("Data de devolução ou ID inválidos. Repita o processo.")


while True:
    
    # menu de opções
    print("################ MENU ################")
    print("1 = CADASTRO ALUNO/LIVRO")
    print("2 = ALTERAÇÃO DE CADASTRO ALUNO/LIVRO")
    print("3 = EMPRÉSTIMO DE LIVRO")
    print("4 = DEVOLUÇÃO DE LIVRO")
    print("5 = RELAÇÃO DE ALUNOS CADASTRADOS")
    print("6 = RELAÇÃO DE LIVROS CADASTRADOS")
    print("7 = RELAÇÃO DE EMPRÉSTIMOS")
    print("0 = ENCERRAR PROGRAMA")
    print("######################################")
    r = input("Escolha a ação: ")
    
    if r == "1":
        print("AREÁ DE CADASTRO")
        while True:
            r = input("Digite [1] para cadastrar Aluno ou [2] para Livro: ")
            if r not in "12":
                print("Esse valor não é válido.")
                continue
            else:
                if r == "1":
                    
                    cadastra_aluno() 
                    while True:               
                        r = input("Deseja cadastrar mais um aluno [S]im [N]ão: ")
                        if r.upper() in "SN":   
                            if r.upper() == "N":
                                break
                            else:
                                cadastra_aluno()
                        else: print("Digite um valor válido.")

                    break
                elif r == "2":
                    
                    cadastra_livro()
                    while True:               
                        r = input("Deseja cadastrar mais um livro [S]im [N]ão: ")
                        if r.upper() in "SN":   
                            if r.upper() == "N":
                                break
                            else:
                                cadastra_livro()
                        else: print("Digite um valor válido.")

                    break 

    elif r == "2":
        while True:
            print("AREÁ DE ALTERAÇÃO DE CADASTRO")
            r = input("Digite [1] para alterar Aluno ou [2] para Livro: ")
            if r not in "12":
                print("Esse valor não é válido.")
                continue
            else:
                if r == "1":
                    altera_aluno() 
                    while True:               
                        r = input("Deseja alterar mais um aluno [S]im [N]ão: ")
                        if r.upper() in "SN":   
                            if r.upper() == "N":
                                break
                            else:
                                altera_aluno()
                        else: print("Digite um valor válido.")

                    break
                else:
                    altera_livro()
                    while True:               
                        r = input("Deseja alterar mais um livro [S]im [N]ão: ")
                        if r.upper() in "SN":   
                            if r.upper() == "N":
                                break
                            else:
                                altera_livro()
                        else: print("Digite um valor válido.")

                    break
        
    elif r == "3":
        print("ÁREA DE EMPRÉSTIMO")
        emprestimo()

    elif r == "4":
        print("ÁREA DE DEVOLUÇÃO")
        devolucao()

    elif r == "5":
        print("RELAÇÃO DE ALUNOS CADASTRADOS")
        print("Gerando...")
        time.sleep(3)
        for item in info_alunos:
            print(info_alunos[item])

    elif r == "6":
        print("RELAÇÃO DE LIVROS CADASTRADOS")
        print("Gerando...")
        time.sleep(3)
        for item in info_livros:
            print(info_livros[item])

    elif r == "7":
        print("RELAÇÃO DE EMPRÉSTIMOS")
        print("Gerando...")
        time.sleep(3)
        for item in dic_emprestimos:
            print(dic_emprestimos[item])

    elif r == "0":
        
        print("Encerrando...")
        database_ids.save("biblioteca.xlsx")
        time.sleep(3)
        break